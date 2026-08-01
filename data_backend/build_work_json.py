#!/usr/bin/env python3
"""Build work-centric JSON chunks from the three source workbooks.

The output follows ``data_schema.md``. Tender links and point coordinates are left
unset unless the source data supports them; the pipeline never fabricates either.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import tempfile
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Iterator
from urllib.parse import quote

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path(__file__).resolve().parent / "csv"
OUTPUT_DIR = Path(__file__).resolve().parent / "json" / "works"
PAYMENTS_FILE = SOURCE_DIR / "2. payments-2026-07-30.xlsx"
JOB_CODES_FILE = SOURCE_DIR / "3. BBMP_job_codes_merged.xlsx"
TENDERS_FILE = SOURCE_DIR / "1. karnataka-tenders-2026-07-30.xlsx"
SOURCE_AS_OF = date(2026, 7, 30)
SCHEMA_VERSION = "0.1"
DEFAULT_MAX_CHUNK_BYTES = 10_000_000
DOCUMENT_BASE_URL = "https://account.bbmpgov.in/vssIFMS/Files1/"
WARD_SCHEMES = {"198", "225", "243", "369", "special"}
WORK_STATUSES = {"in_progress", "completed", "unknown"}
LINK_FAILURE_REASONS = {
    "no_attachment",
    "blank_document",
    "mislabelled_document",
    "unreadable",
    "number_not_in_corpus",
    "no_candidate_found",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def nullable_text(value: Any) -> str | None:
    text = clean_text(value)
    return text or None


def money_to_int(value: Any) -> int:
    """Round source rupee decimals to the nearest rupee, half up."""
    if value in (None, ""):
        return 0
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = clean_text(value)
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        except ValueError:
            return None
    # 1900-01-01 is used as a sentinel for missing dates in the source.
    return None if parsed.year <= 1900 else parsed.isoformat()


def award_fy(job_number: str) -> str:
    # Most IDs are ``ward-YY-serial``. Central and migrated records retain the
    # same trailing FY/serial structure but add prefixes such as ``R-`` or
    # ``BlockCAO``; the source schema still treats the complete string as the key.
    match = re.search(r"-(\d{2})-\d{6}$", job_number)
    if not match:
        raise ValueError(f"Invalid job number: {job_number!r}")
    start_short = int(match.group(1))
    start_year = 2000 + start_short
    return f"{start_year:04d}-{(start_short + 1) % 100:02d}"


def contractor_key(name: str) -> str:
    value = unicodedata.normalize("NFKD", name).casefold()
    value = re.sub(r"^\s*\d{4,8}\s+", "", value)
    value = re.sub(r"^\s*m\s*[/.\\-]?\s*s\s*[/.\\-]?\s*", "", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def contractor_display_name(name: str) -> str:
    """Remove the internal vendor-code prefix from a display name."""
    return re.sub(r"^\s*\d{4,8}\s+", "", clean_text(name)).strip()


def is_placeholder(filename: str) -> bool:
    value = unicodedata.normalize("NFKD", filename).casefold()
    value = re.sub(r"[_-]+", " ", value)
    return bool(
        re.search(r"\bblank(?:\s*sheet)?\b", value)
        or re.search(r"\bnot\s+app(?:licable|licale|licalbe|icable|icabl)\b", value)
    )


def worksheet_rows(path: Path) -> Iterator[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["in"]
        rows = worksheet.iter_rows(values_only=True)
        headers = tuple(next(rows))
        if not all(isinstance(header, str) for header in headers):
            raise ValueError(f"{path}: header row contains a non-string value")
        for values in rows:
            yield dict(zip(headers, values, strict=True))
    finally:
        workbook.close()


def parse_documents(value: Any) -> Iterator[tuple[str, str]]:
    if value in (None, ""):
        return
    for item in str(value).split(" | "):
        if ": " not in item:
            raise ValueError(f"Malformed document entry: {item!r}")
        doc_type, filename = item.split(": ", 1)
        doc_type = clean_text(doc_type)
        filename = clean_text(filename)
        if not doc_type or not filename:
            raise ValueError(f"Malformed document entry: {item!r}")
        yield doc_type, filename


def bill_identity(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        clean_text(row["WO"]),
        clean_text(row["billtype"]),
        clean_text(row["commonbrnumber"]),
        iso_date(row["commonbrdate"]),
        clean_text(row["rtgs"]),
        iso_date(row["rtgsdate"]),
        contractor_key(clean_text(row["contractorname"])),
        money_to_int(row["gross"]),
        money_to_int(row["nett"]),
        money_to_int(row["deduction"]),
    )


def choose_weighted(values: dict[str, list[int]]) -> str:
    if not values:
        return ""
    return max(
        values,
        key=lambda value: (
            sum(abs(amount) for amount in values[value]),
            len(values[value]),
            len(value),
            value.casefold(),
        ),
    )


def normalize_scheme(value: Any) -> str:
    if value in (None, ""):
        return "special"
    if isinstance(value, (int, float)) and float(value).is_integer():
        scheme = str(int(value))
    else:
        scheme = clean_text(value).casefold()
    return scheme if scheme in WARD_SCHEMES else "special"


def description_similarity(left: str, right: str) -> float:
    def normalized(value: str) -> str:
        value = unicodedata.normalize("NFKD", value).casefold()
        return " ".join(re.findall(r"[a-z0-9]+", value))

    return SequenceMatcher(None, normalized(left), normalized(right)).ratio()


def choose_payment_metadata(
    description: str, candidates: list[dict[str, Any]]
) -> dict[str, Any] | None:
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda row: (
            description_similarity(description, clean_text(row["work_description"])),
            clean_text(row["work_description"]) == description,
            clean_text(row["ward_name"]),
        ),
    )


def load_payment_metadata(target_jobs: set[str]) -> dict[str, list[dict[str, Any]]]:
    candidates: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen: set[tuple[Any, ...]] = set()
    fields = (
        "job_number",
        "work_description",
        "ward_number",
        "ward_name",
        "ward_scheme",
        "zone",
        "budget_head",
    )
    for row in worksheet_rows(PAYMENTS_FILE):
        job_number = clean_text(row["job_number"])
        if job_number not in target_jobs:
            continue
        key = tuple(row[field] for field in fields)
        if key in seen:
            continue
        seen.add(key)
        candidates[job_number].append({field: row[field] for field in fields})
    return candidates


def collect_works() -> tuple[dict[str, dict[str, Any]], dict[str, int]]:
    works: dict[str, dict[str, Any]] = {}
    raw_rows = 0
    duplicate_bill_rows = 0

    for row in worksheet_rows(JOB_CODES_FILE):
        raw_rows += 1
        job_number = clean_text(row["WO"])
        if not job_number:
            raise ValueError("Job-code row has no WO")
        work = works.setdefault(
            job_number,
            {
                "bill_keys": set(),
                "bills": [],
                "documents": set(),
                "description_amounts": defaultdict(list),
                "contractor_amounts": defaultdict(list),
            },
        )

        for document in parse_documents(row["documents_detail"]):
            work["documents"].add(document)

        identity = bill_identity(row)
        if identity in work["bill_keys"]:
            duplicate_bill_rows += 1
            continue
        work["bill_keys"].add(identity)

        gross = money_to_int(row["gross"])
        net = money_to_int(row["nett"])
        deduction = money_to_int(row["deduction"])
        if gross - deduction != net:
            raise ValueError(
                f"{job_number}: source bill violates gross - deduction == net"
            )

        description = clean_text(row["wcdescription"])
        contractor = contractor_display_name(row["contractorname"])
        work["description_amounts"][description].append(gross)
        work["contractor_amounts"][contractor].append(gross)
        work["bills"].append(
            {
                "bill_type": clean_text(row["billtype"]),
                "gross": gross,
                "net": net,
                "deduction": deduction,
                "bill_date": iso_date(row["commonbrdate"]),
                "rtgs_date": iso_date(row["rtgsdate"]),
            }
        )

    stats = {
        "source_rows": raw_rows,
        "deduplicated_bill_rows": duplicate_bill_rows,
        "work_count": len(works),
        "work_description_conflicts": sum(
            len(work["description_amounts"]) > 1 for work in works.values()
        ),
        "work_contractor_conflicts": sum(
            len(work["contractor_amounts"]) > 1 for work in works.values()
        ),
    }
    return works, stats


def count_tender_corpus() -> int:
    count = 0
    for row in worksheet_rows(TENDERS_FILE):
        if not clean_text(row.get("tenderNumber")):
            raise ValueError("Tender corpus row has no tenderNumber")
        count += 1
    return count


def document_objects(documents: set[tuple[str, str]]) -> list[dict[str, Any]]:
    return [
        {
            "doc_type": doc_type,
            "filename": filename,
            "url": DOCUMENT_BASE_URL + quote(filename, safe=""),
            "is_placeholder": is_placeholder(filename),
        }
        for doc_type, filename in sorted(
            documents, key=lambda item: (item[0].casefold(), item[1].casefold())
        )
    ]


def tender_failure_reason(documents: list[dict[str, Any]]) -> str:
    tender_documents = [
        document for document in documents if document["doc_type"] == "Tender Documents"
    ]
    if not tender_documents:
        return "no_attachment"
    if all(document["is_placeholder"] for document in tender_documents):
        return "blank_document"
    # The structured workbooks contain no KPPP tender number. A real attachment is
    # evidence that matching may be possible later, but not enough to invent a link.
    return "no_candidate_found"


def current_award_fy(as_of: date) -> str:
    start_year = as_of.year if as_of.month >= 4 else as_of.year - 1
    return f"{start_year:04d}-{(start_year + 1) % 100:02d}"


def build_location(
    scheme: str,
    ward_name: str | None,
    zone: str | None,
) -> dict[str, Any] | None:
    if scheme == "special":
        return {
            "lat": None,
            "lng": None,
            "precision": "none",
            "source": "ward_centroid",
            "confidence": "low",
            "place_name": zone,
            "notes": "Centralised or non-ward work; no point coordinates in source CSVs.",
        }
    if not ward_name:
        return None
    return {
        "lat": None,
        "lng": None,
        "precision": "ward",
        "source": "ward_centroid",
        "confidence": "low",
        "place_name": f"{ward_name}, Bengaluru",
        "notes": "Ward-level location only; no point coordinates in source CSVs.",
    }


def assemble_record(
    job_number: str,
    work: dict[str, Any],
    payment_candidates: list[dict[str, Any]],
    as_of: date,
) -> dict[str, Any]:
    description = choose_weighted(work["description_amounts"])
    contractor_name = choose_weighted(work["contractor_amounts"])
    metadata = choose_payment_metadata(description, payment_candidates)

    if metadata is None:
        scheme = "special"
        ward_number = None
        ward_name = None
        zone = None
        budget_head = None
    else:
        scheme = normalize_scheme(metadata["ward_scheme"])
        if scheme == "special":
            ward_number = None
            ward_name = None
        else:
            ward_number = (
                int(metadata["ward_number"])
                if metadata["ward_number"] not in (None, "")
                else None
            )
            ward_name = nullable_text(metadata["ward_name"])
        zone = nullable_text(metadata["zone"])
        budget_head = nullable_text(metadata["budget_head"])

    bills = sorted(
        work["bills"],
        key=lambda bill: (
            bill["bill_date"] or "9999-12-31",
            bill["rtgs_date"] or "9999-12-31",
            bill["bill_type"],
            bill["gross"],
        ),
    )
    gross = sum(bill["gross"] for bill in bills)
    net = sum(bill["net"] for bill in bills)
    deduction = sum(bill["deduction"] for bill in bills)
    bill_dates = [bill["bill_date"] for bill in bills if bill["bill_date"]]
    documents = document_objects(work["documents"])
    completion_documents = [
        document
        for document in documents
        if document["doc_type"] == "Completion Certificate"
        and not document["is_placeholder"]
    ]
    fy = award_fy(job_number)
    if completion_documents:
        status = "completed"
    elif fy == current_award_fy(as_of):
        status = "in_progress"
    else:
        status = "unknown"

    return {
        "job_number": job_number,
        "description": description,
        "ward_number": ward_number,
        "ward_name": ward_name,
        "ward_scheme": scheme,
        "zone": zone,
        "award_fy": fy,
        "contractor_name": contractor_name,
        "contractor_key": contractor_key(contractor_name),
        "budget_head": budget_head,
        "amount_gross": gross,
        "amount_net": net,
        "amount_deduction": deduction,
        "estimated_amount": None,
        "bill_count": len(bills),
        "first_bill_date": min(bill_dates) if bill_dates else None,
        "last_bill_date": max(bill_dates) if bill_dates else None,
        "status": status,
        "tender": None,
        "link_failure_reason": tender_failure_reason(documents),
        "location": build_location(scheme, ward_name, zone),
        "documents": documents,
        "bills": bills,
    }


def validate_record(record: dict[str, Any]) -> None:
    required = {
        "job_number",
        "description",
        "ward_number",
        "ward_name",
        "ward_scheme",
        "zone",
        "award_fy",
        "contractor_name",
        "contractor_key",
        "budget_head",
        "amount_gross",
        "amount_net",
        "amount_deduction",
        "estimated_amount",
        "bill_count",
        "first_bill_date",
        "last_bill_date",
        "status",
        "tender",
        "location",
        "documents",
        "bills",
    }
    missing = required - record.keys()
    if missing:
        raise ValueError(f"{record.get('job_number')}: missing fields {sorted(missing)}")
    if award_fy(record["job_number"]) != record["award_fy"]:
        raise ValueError(f"{record['job_number']}: invalid award_fy")
    if not record["description"] or not record["contractor_name"]:
        raise ValueError(f"{record['job_number']}: missing description or contractor")
    if contractor_key(record["contractor_name"]) != record["contractor_key"]:
        raise ValueError(f"{record['job_number']}: contractor_key mismatch")
    if record["ward_scheme"] not in WARD_SCHEMES:
        raise ValueError(f"{record['job_number']}: invalid ward_scheme")
    if record["ward_number"] is not None and not isinstance(record["ward_number"], int):
        raise ValueError(f"{record['job_number']}: non-integer ward_number")
    if record["ward_number"] is not None and record["ward_scheme"] == "special":
        raise ValueError(f"{record['job_number']}: ward number without ward scheme")
    if record["status"] not in WORK_STATUSES:
        raise ValueError(f"{record['job_number']}: invalid status")
    for field in ("amount_gross", "amount_net", "amount_deduction"):
        if not isinstance(record[field], int):
            raise ValueError(f"{record['job_number']}: {field} is not an integer")
    if record["estimated_amount"] is not None and not isinstance(
        record["estimated_amount"], int
    ):
        raise ValueError(f"{record['job_number']}: estimated_amount is not an integer")
    if record["amount_gross"] - record["amount_deduction"] != record["amount_net"]:
        raise ValueError(f"{record['job_number']}: aggregate amount invariant failed")
    if record["bill_count"] != len(record["bills"]):
        raise ValueError(f"{record['job_number']}: bill_count mismatch")
    for bill in record["bills"]:
        if set(bill) != {"bill_type", "gross", "net", "deduction", "bill_date", "rtgs_date"}:
            raise ValueError(f"{record['job_number']}: invalid bill fields")
        if bill["gross"] - bill["deduction"] != bill["net"]:
            raise ValueError(f"{record['job_number']}: bill amount invariant failed")
        if not all(isinstance(bill[field], int) for field in ("gross", "net", "deduction")):
            raise ValueError(f"{record['job_number']}: non-integer bill amount")
        for field in ("bill_date", "rtgs_date"):
            if bill[field] is not None:
                date.fromisoformat(bill[field])
    for document in record["documents"]:
        if set(document) != {"doc_type", "filename", "url", "is_placeholder"}:
            raise ValueError(f"{record['job_number']}: invalid document fields")
        expected_url = DOCUMENT_BASE_URL + quote(document["filename"], safe="")
        if document["url"] != expected_url or not isinstance(document["is_placeholder"], bool):
            raise ValueError(f"{record['job_number']}: invalid document metadata")
    if record["tender"] is None:
        reason = record.get("link_failure_reason")
        if reason not in LINK_FAILURE_REASONS:
            raise ValueError(f"{record['job_number']}: missing tender failure reason")
    else:
        if not {"link_method", "link_confidence"} <= record["tender"].keys():
            raise ValueError(f"{record['job_number']}: unqualified tender link")
        contract = record["tender"].get("contract")
        if contract and contract.get("job_number") not in (None, record["job_number"]):
            raise ValueError(f"{record['job_number']}: contract job number mismatch")
    location = record["location"]
    if location:
        if location["precision"] not in {"point", "ward", "zone", "none"}:
            raise ValueError(f"{record['job_number']}: invalid location precision")
        if location["source"] not in {
            "photo_overlay",
            "geocoded_text",
            "ward_centroid",
            "manual",
        }:
            raise ValueError(f"{record['job_number']}: invalid location source")
        if location["confidence"] not in {"high", "medium", "low"}:
            raise ValueError(f"{record['job_number']}: invalid location confidence")
    if location and location["precision"] == "point":
        if location["lat"] is None or location["lng"] is None:
            raise ValueError(f"{record['job_number']}: point location lacks coordinates")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def manifest_path_for(output_dir: Path) -> Path:
    return output_dir.with_name(f"{output_dir.name}.manifest.json")


def write_output(
    records: Iterable[dict[str, Any]],
    output_dir: Path,
    max_chunk_bytes: int,
    build_stats: dict[str, int],
) -> dict[str, Any]:
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    stats = Counter()
    seen_jobs: set[str] = set()
    temporary_dir = Path(
        tempfile.mkdtemp(prefix="works-json-", dir=output_dir.parent)
    )
    chunks: list[dict[str, Any]] = []
    serialized_records: list[str] = []
    serialized_bytes = 3  # Opening bracket, closing bracket, trailing newline.

    def flush_chunk() -> None:
        nonlocal serialized_bytes
        if not serialized_records:
            return
        filename = f"works-{len(chunks) + 1:04d}.json"
        chunk_path = temporary_dir / filename
        with chunk_path.open("w", encoding="utf-8") as destination:
            destination.write("[")
            destination.write(",".join(serialized_records))
            destination.write("]\n")
        chunk_size = chunk_path.stat().st_size
        if chunk_size > max_chunk_bytes:
            raise ValueError(
                f"{filename} is {chunk_size:,} bytes; limit is {max_chunk_bytes:,}"
            )
        chunks.append(
            {
                "file": f"{output_dir.name}/{filename}",
                "records": len(serialized_records),
                "bytes": chunk_size,
                "sha256": sha256(chunk_path),
            }
        )
        serialized_records.clear()
        serialized_bytes = 3

    try:
        for record in records:
            validate_record(record)
            if record["job_number"] in seen_jobs:
                raise ValueError(f"Duplicate job number: {record['job_number']}")
            encoded = json.dumps(
                record,
                ensure_ascii=False,
                separators=(",", ":"),
            )
            encoded_bytes = len(encoded.encode("utf-8"))
            separator_bytes = 1 if serialized_records else 0
            if encoded_bytes + 3 > max_chunk_bytes:
                raise ValueError(
                    f"{record['job_number']} cannot fit in a "
                    f"{max_chunk_bytes:,}-byte chunk"
                )
            if serialized_records and (
                serialized_bytes + separator_bytes + encoded_bytes > max_chunk_bytes
            ):
                flush_chunk()
                separator_bytes = 0
            serialized_records.append(encoded)
            serialized_bytes += separator_bytes + encoded_bytes

            seen_jobs.add(record["job_number"])
            stats[f"status_{record['status']}"] += 1
            stats[f"tender_failure_{record['link_failure_reason']}"] += 1
            stats["documents"] += len(record["documents"])
            stats["bills"] += len(record["bills"])
            stats["placeholders"] += sum(
                document["is_placeholder"] for document in record["documents"]
            )
        flush_chunk()

        backup_dir = output_dir.with_name(f".{output_dir.name}-previous")
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
        if output_dir.exists():
            output_dir.replace(backup_dir)
        temporary_dir.replace(output_dir)
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
    except BaseException:
        shutil.rmtree(temporary_dir, ignore_errors=True)
        raise

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "schema_file": "../../data_schema.md",
        "as_of_date": SOURCE_AS_OF.isoformat(),
        "sources": [
            str(TENDERS_FILE.relative_to(REPO_ROOT)),
            str(PAYMENTS_FILE.relative_to(REPO_ROOT)),
            str(JOB_CODES_FILE.relative_to(REPO_ROOT)),
        ],
        "record_count": len(seen_jobs),
        "chunking": {
            "max_bytes": max_chunk_bytes,
            "chunk_count": len(chunks),
            "total_bytes": sum(chunk["bytes"] for chunk in chunks),
        },
        "chunks": chunks,
        "stats": {**build_stats, **dict(sorted(stats.items()))},
        "methodology": {
            "root": "One record per distinct WO in BBMP_job_codes_merged.xlsx.",
            "bills": (
                "Repeated snapshot rows are deduplicated by work, bill references, "
                "dates, contractor, type, and amounts before aggregation."
            ),
            "work_conflicts": (
                "When a WO has multiple source descriptions or contractors, all "
                "distinct bills remain in the work totals; the description and "
                "contractor attached to the largest gross total become canonical."
            ),
            "ward_metadata": (
                "Payments rows are matched on job_number; scheme collisions are "
                "resolved by similarity to the canonical description. Special "
                "schemes emit no ward number or name."
            ),
            "tenders": (
                "No structured source field links a WO to a full KPPP tender number, "
                "so tender is null and link_failure_reason is explicit."
            ),
            "locations": (
                "Source CSVs have no coordinates; locations are ward-level or "
                "precision none."
            ),
            "status": (
                "A non-placeholder Completion Certificate means completed. "
                "Current-FY work without one is in_progress; all other cases are "
                "unknown."
            ),
            "amounts": (
                "Decimal source amounts are rounded half-up to integer rupees; "
                "source and aggregate accounting identities are asserted."
            ),
        },
    }
    manifest_path = manifest_path_for(output_dir)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        prefix="works-manifest-",
        suffix=".tmp",
        dir=output_dir.parent,
        delete=False,
    ) as destination:
        temporary_manifest = Path(destination.name)
        json.dump(manifest, destination, ensure_ascii=False, indent=2)
        destination.write("\n")
    temporary_manifest.replace(manifest_path)
    manifest_path.chmod(0o644)
    return manifest


def validate_output(output_dir: Path) -> dict[str, Any]:
    manifest_path = manifest_path_for(output_dir)
    with manifest_path.open(encoding="utf-8") as source:
        manifest = json.load(source)
    seen: set[str] = set()
    stats = Counter()
    total_bytes = 0
    max_chunk_bytes = manifest["chunking"]["max_bytes"]
    if manifest["chunking"]["chunk_count"] != len(manifest["chunks"]):
        raise ValueError("Manifest chunk_count mismatch")
    for chunk in manifest["chunks"]:
        chunk_path = manifest_path.parent / chunk["file"]
        chunk_size = chunk_path.stat().st_size
        if chunk_size > max_chunk_bytes:
            raise ValueError(f"Chunk exceeds size limit: {chunk_path}")
        if chunk_size != chunk["bytes"] or sha256(chunk_path) != chunk["sha256"]:
            raise ValueError(f"Checksum or size mismatch: {chunk_path}")
        with chunk_path.open(encoding="utf-8") as source:
            records = json.load(source)
        if not isinstance(records, list) or len(records) != chunk["records"]:
            raise ValueError(f"Record count mismatch: {chunk_path}")
        total_bytes += chunk_size
        for record in records:
            validate_record(record)
            job_number = record["job_number"]
            if job_number in seen:
                raise ValueError(f"Duplicate job number: {job_number}")
            seen.add(job_number)
            stats[f"status_{record['status']}"] += 1
            stats[f"tender_failure_{record['link_failure_reason']}"] += 1
            stats["documents"] += len(record["documents"])
            stats["bills"] += len(record["bills"])
            stats["placeholders"] += sum(
                document["is_placeholder"] for document in record["documents"]
            )
    if total_bytes != manifest["chunking"]["total_bytes"]:
        raise ValueError("Manifest total_bytes mismatch")
    if len(seen) != manifest["record_count"]:
        raise ValueError("Manifest record_count mismatch")
    for key, value in stats.items():
        if manifest["stats"].get(key) != value:
            raise ValueError(f"Manifest stats mismatch for {key}")
    return manifest


def build(output_dir: Path, max_chunk_bytes: int) -> dict[str, Any]:
    works, stats = collect_works()
    payment_metadata = load_payment_metadata(set(works))
    stats["tender_corpus_rows"] = count_tender_corpus()
    stats["works_with_payment_metadata"] = len(payment_metadata)
    stats["works_without_payment_metadata"] = len(works) - len(payment_metadata)
    stats["payment_metadata_conflicts"] = sum(
        len(candidates) > 1 for candidates in payment_metadata.values()
    )
    records = (
        assemble_record(
            job_number,
            works[job_number],
            payment_metadata.get(job_number, []),
            SOURCE_AS_OF,
        )
        for job_number in sorted(works)
    )
    return write_output(records, output_dir, max_chunk_bytes, stats)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument(
        "--max-chunk-bytes",
        type=int,
        default=DEFAULT_MAX_CHUNK_BYTES,
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Validate existing output instead of rebuilding it.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.max_chunk_bytes <= 3:
        raise SystemExit("--max-chunk-bytes must be greater than 3")
    manifest = (
        validate_output(args.output_dir)
        if args.check
        else build(args.output_dir, args.max_chunk_bytes)
    )
    print(
        f"Validated {manifest['record_count']:,} work records in "
        f"{manifest['chunking']['chunk_count']} chunk(s)."
    )


if __name__ == "__main__":
    main()
